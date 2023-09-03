from app import log
from app.data import student_intake as mstudent, settings as msettings, utils as mutils
import app.data.settings
from app.application import formio as mformio
import sys, datetime, requests, json
from openpyxl import load_workbook
from io import BytesIO


toolbox2db_keys = {
    "personalia_rijksregisternummer": "s_rijksregister",
    "personalia_achternaam": "s_last_name",
    "personalia_voornaam": "s_first_name",
    "personalia_roepnaam": "s_roepnaam",
    "personalia_geboorteplaats": "s_geboorteplaats",
    "personalia_geboorteland": "s_geboorteland",
    "personalia_nationaliteit": "s_nationaliteit",
    "personalia_geslacht": "s_sex",
    "personalia_geboortedatum": "s_date_of_birth",
    "Taalachtergrond_Thuistaal Nederland": "g_spreektaal_Nederland",
    "Taalachtergrond_Andere talen": "g_spreektaal_Ander",
    "Taalachtergrond_Andere thuistaal": "g_spreektaal_Thuistaal",
    "Taalachtergrond_Nederlandstalig oudercontact": "g_spreektaal_Oudercontact",
    "Gezinssituatie_Aantal kinderen": "g_aantal_kinderen",
    "SCHOOLLOOPBAAN_Naam (basis)school": "school_naam",
    "SCHOOLLOOPBAAN_Type BLO": "school_blo_type",
    "SCHOOLLOOPBAAN_Gemeente (basis)school": "school_adres",
    "SCHOOLLOOPBAAN_(Voorlopig) advies basisschool": "school_voorlopig_advies",
    "SCHOOLLOOPBAAN_Extra informatie": "school_extra_info",
    "SCHOOLLOOPBAAN_CLIL": "clil_keuze",
    "Speciale zorg_Zorgattest beschikbaar": "f_zorgattest_beschikbaar",
    "Speciale zorg_Intakegesprek": "f_intakegesprek_gewenst",
    "Speciale zorg_Bijkomende opmerkingen": "extra_begeleiding_welke",
    "Speciale zorg_Begeleiding": "extra_begeleiding_door_wie",
    "Speciale zorg_Ondersteuningsnetwerk": "f_ondersteuningsnetwerk",
    "Speciale zorg_Specifieke onderwijsbehoefte(n)": "specifieke_onderwijsbehoeften",
    "Speciale zorg_Verhoogde zorg": "andere_schoolproblemen",
    "Speciale zorg_Gezondheidsproblemen": "g_problemen",
    "Speciale zorg_Medicatie": "g_medicatie",
    "Speciale zorg_Naam huisarts": "g_huisarts_naam",
    "Speciale zorg_Gemeente huisarts": "g_huisarts_adres",
    "Betalingen_e-mailadres": "betaling_email",
    "Betalingen_Betalingen": "betaling_wijze",
    "Betalingen_IBAN": "domiciliering_iban",
    "Betalingen_BIC": "domiciliering_bic",
    "Betalingen_Rekeninghouder": "domiciliering_rekeninghouder",
}

TOPIC_IN_ROW = 1
HEADER_IN_ROW = 2

def XLSXDictReader(f):
    book = load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=HEADER_IN_ROW, column=i).value) for i in range(1, cols))
    topics = [""]
    if TOPIC_IN_ROW > -1:
        topic = ""
        for i in range(1, cols + 1):
            value = sheet.cell(row=TOPIC_IN_ROW, column=i).value
            if value and value != topic:
                topic = value
            topics.append(topic + "_")
    else:
        topics = ["" for i in range(1, cols + 1)]

    def item(i, j):
        value = sheet.cell(row=i, column=j).value
        header = sheet.cell(row=HEADER_IN_ROW, column=j).value
        header = topics[j] + header if header else None
        if isinstance(value, str):
            value = value.strip()
        if not header :
            return ("", value)
        if header in toolbox2db_keys:
            return (toolbox2db_keys[header], value)
        else:
            return (header, value)

    return (dict(item(i, j) for j in range(1, cols + 1)) for i in range(HEADER_IN_ROW + 1, rows + 1))


def import_student_info(file_storeage):
    try:
        now = datetime.datetime.now()
        db_students = mstudent.get_students({"enabled": True})
        db_students_cache = {s.s_last_name+s.s_first_name+str(s.s_date_of_birth): s for s in db_students }
        toolbox_students = XLSXDictReader(BytesIO(file_storeage.read()))
        for tb_student in toolbox_students:
            tb_student["g_spreektaal"] = tb_student["g_spreektaal_Ander"] if tb_student["g_spreektaal_Ander"] != "/" else "NL"
            tb_student["i_intake_date"] = now.strftime('%d/%m/%Y %H:%M')
            tb_student["i_last_name"] = "import"
            tb_key = tb_student["s_last_name"] + tb_student["s_first_name"] + str(mformio.datestring_to_date(tb_student['s_date_of_birth']))
            if tb_key in db_students_cache:
                db_student = db_students_cache[tb_key]
                tb_student["id"] = db_student.id
                update_student(tb_student)
            else:
                add_student(tb_student)
    except Exception as e:
        mutils.raise_error(f'{sys._getframe().f_code.co_name}:', e)
    return None



def add_student(data):
    try:
        data['s_date_of_birth'] = mformio.datestring_to_date(data['s_date_of_birth'])
        data['i_intake_date'] = mformio.datetimestring_to_datetime(data['i_intake_date'])
        student = mstudent.add_student(data)
        log.info(f"Add intake: {student.s_last_name} {student.s_first_name}, {data}")
        return {"status": True, "data": {'id': student.id}}
    except Exception as e:
        log.error(f'{sys._getframe().f_code.co_name}: {e}')
        log.error(data)
        return {"status": False, "data": f'generic error {e}'}


def update_student(data):
    try:
        student = mstudent.get_first_student({'id': data['id']})
        if student:
            del data['id']
            data['s_date_of_birth'] = mformio.datestring_to_date(data['s_date_of_birth'])
            data['i_intake_date'] = mformio.datetimestring_to_datetime(data['i_intake_date'])
            student = mstudent.update_student(student, data)
            if student:
                log.info(f"Update intake: {student.s_last_name} {student.s_first_name}, {data}")
                return {"status": True, "data": {'id': student.id}}
        return {"status": False, "data": "Er is iets fout gegaan"}
    except Exception as e:
        log.error(f'{sys._getframe().f_code.co_name}: {e}')
        log.error(data)
        return {"status": False, "data": f'generic error {e}'}


def delete_students(ids):
    mstudent.delete_students(ids)


def get_unique_klassen():
    return mstudent.get_unique_klassen()

############## formio forms #############
def prepare_add_form():
    try:
        template = app.data.settings.get_json_template('intake-formio-template')
        now = datetime.datetime.now()
        return {'template': template,
                'defaults': {
                        'i_intake_date': mformio.datetime_to_datetimestring(now),
                        's_code': msettings.get_and_increment_default_student_code()
                }
                }
    except Exception as e:
        log.error(f'{sys._getframe().f_code.co_name}: {e}')
        raise e


def prepare_edit_form(id):
    try:
        student = mstudent.get_first_student({"id": id})
        template = app.data.settings.get_json_template('intake-formio-template')
        template = mformio.prepare_for_edit(template, student.to_dict())
        return {'template': template,
                'defaults': student.to_dict()}
    except Exception as e:
        log.error(f'{sys._getframe().f_code.co_name}: {e}')
        raise e


############ care overview list #########
def format_data(db_list):
    out = []
    for student in db_list:
        em = student.to_dict()
        em.update({
            'row_action': student.id,
            'DT_RowId': student.id
        })
        if student.klas == '':
            em.update({'overwrite_row_color': "rgb(232, 182, 120)"})
        out.append(em)
    return out


################  Cron jobs ##################

# get list of students from sdh
# use rijksregisternummer OR naam+voornaam as key to find a student
# fill in the classnumber and wisa internal number
def link_students_to_class_cron_task(opaque):
    try:
        if msettings.get_configuration_setting('cron-enable-update-student-class'):
            sdh_url = msettings.get_configuration_setting('sdh-base-url')
            sdh_key = msettings.get_configuration_setting('sdh-api-key')
            session = requests.Session()
            res = session.get(sdh_url, headers={'x-api-key': sdh_key})
            if res.status_code == 200:
                sdh_students = res.json()
                nbr_student_matching_rijkregister_found = 0
                nbr_student_matching_naam_found = 0
                nbr_student_not_found = 0
                if sdh_students['status']:
                    log.info(f'{sys._getframe().f_code.co_name}, retrieved {len(sdh_students["data"])} students from SDH')
                    rijksregister_to_student = {s['rijksregisternummer']: s for s in sdh_students["data"]}
                    naam_to_student = {s['naam'] + s['voornaam']: s for s in sdh_students["data"]}
                    students = mstudent.get_students()
                    log.info(f'{sys._getframe().f_code.co_name}, {len(students)} students in database')
                    for student in students:
                        name = student.s_last_name + student.s_first_name
                        rijksregisternummer = student.s_rijksregister.replace('-', '').replace('.', '')
                        if rijksregisternummer != "" and rijksregisternummer in rijksregister_to_student:
                            nbr_student_matching_rijkregister_found += 1
                            student.klas = rijksregister_to_student[rijksregisternummer]['klascode']
                            student.s_code = rijksregister_to_student[rijksregisternummer]['leerlingnummer']
                        elif name in naam_to_student:
                            nbr_student_matching_naam_found += 1
                            student.klas = naam_to_student[name]['klascode']
                            student.s_code = naam_to_student[name]['leerlingnummer']
                        else:
                            nbr_student_not_found += 1
                    mstudent.commit()
                    log.info(f'{sys._getframe().f_code.co_name}, students, matching rijksregisternummer, found {nbr_student_matching_rijkregister_found}')
                    log.info(f'{sys._getframe().f_code.co_name}, students, matching naam, found {nbr_student_matching_naam_found}')
                    log.info(f'{sys._getframe().f_code.co_name}, students not found {nbr_student_not_found}')
                else:
                    log.error(f'{sys._getframe().f_code.co_name}: sdh returned {sdh_students["data"]}')

            else:
                log.error(f'{sys._getframe().f_code.co_name}: api call returned {res.status_code}')

    except Exception as e:
        log.error(f'{sys._getframe().f_code.co_name}: {e}')
